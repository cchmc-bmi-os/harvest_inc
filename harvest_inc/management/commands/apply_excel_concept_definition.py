"""
## Apply modifications specified

## RUN IT LIKE THIS:
## - Save current state of auth, category, concepts and fields (in case it needs to be re-run)
## - if re-running:
##    - TRUNCATE cascage categories (if re-running(
##    - loaddata fixture.json
## - ibemc/etl/6-all-elements/ETL_revised_concepts_20170426.py | bin/manage.py shell

"""
from optparse import make_option
from django.core.management.base import BaseCommand
from avocado.models import DataCategory, DataConcept
from harvest_tools.tools import update_object as uo
from openpyxl import load_workbook

## file_name = 'ibemc/etl/7-cde-pku-mcad/clean_MCAD_specs.xlsx'

def apply_concept_definition(filename):
    wb = load_workbook(filename=filename)

    ## Orders redefines Categories and Concepts
    ws = wb.active
    print(filename)
    print(ws)
    
    order = 0
    categories = []
    for row in ws.iter_rows():
        print(row)
        order = order + 1
        print(order)
        category = row[0].value
        sub_category = row[1].value
        current_concept = row[2].value
        revised_concept = row[3].value
        published = True if row[4].value in ['t', 'T', 'True', 'TRUE'] else False
            
        if category != 'Revised Category':
            print('published', published)
            categories.append(category)
            print('Checking {0}'.format(category))
            uo(category, DataCategory, {'published': True, 'parent_category': None, 'order': order}, False, True, 'name')
            if sub_category:
                print('Checking {0}'.format(sub_category))
                categories.append(sub_category)
                uo(sub_category, DataCategory, {'published': True, 'parent_category': category, 'order': order}, False, True, 'name')
                
            new_category = sub_category if sub_category else category
            uo(current_concept, DataConcept, {'name': revised_concept, 'published': published, 'category_name': new_category, 'order': order}, False, True, 'id')


class Command(BaseCommand):
    """Apply concept definition

    """
    help = 'Apply concept definition'

    option_list = BaseCommand.option_list + (
        make_option(
            '--filename',
            dest='filename',
            help='Excel file path',
        ),
    )

    def handle(self, *args, **options):
        print('Start...')
        filename = options['filename']
        apply_concept_definition(filename)
        
        print('Done.')
